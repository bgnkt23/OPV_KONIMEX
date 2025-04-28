from openpyxl import load_workbook

def extract_multi_level_headers(excel_file, start_row=4, num_levels=3):
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    headers = []
    max_col = ws.max_column

    # Fungsi untuk menyederhanakan header utama
    def simplify_main_header(header_text):
        if "-" in header_text:
            return header_text.split("-")[0].strip()
        return header_text.strip()

    # Buat list header per kolom
    for col in range(1, max_col + 1):
        levels = []
        for row in range(num_levels):
            row = start_row + row_offset
            cell = ws.cell(row=row, column=col)

            # Deteksi merge cell
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break

            value = str(cell.value).strip() if cell.value else ""
            col_header.append(value)

        # Kolom A, B, C khusus ambil 1 level pertama
        if col <= 3:
            headers.append(levels[0])
        else:
            # Header utama disingkat
            simplified_main = simplify_main_header(levels[0])
            combined = " > ".join([simplified_main] + levels[1:])
            headers.append(combined)

    return headers
